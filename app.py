# sistema_contador/app.py
from flask import Flask, render_template, request, jsonify, g, send_file
from database import get_db_connection, init_database
from modelos_ia import CadenaMarkovPagos
from evaluacion_credito import EvaluadorCredito
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import io
import base64
import random
from dotenv import load_dotenv
import os
import webbrowser
import threading
import time
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import openpyxl

# Cargar variables de entorno
load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'clave_secreta_sistema_contador')

# Contexto global para variables compartidas
@app.before_request
def cargar_variables_globales():
    """Cargar variables globales para todas las templates"""
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor(dictionary=True)
            
            # Total de clientes
            cursor.execute("SELECT COUNT(*) as total FROM clientes")
            g.total_clientes = cursor.fetchone()['total']
            
            # Clientes en riesgo (con impagos)
            cursor.execute("""
                SELECT COUNT(DISTINCT c.id) as count 
                FROM clientes c
                LEFT JOIN pagos p ON c.id = p.cliente_id
                WHERE p.estado = 'impago' OR c.estado_actual = 'impago'
            """)
            g.clientes_riesgo_count = cursor.fetchone()['count']
            
            # Pagos pendientes
            cursor.execute("""
                SELECT COUNT(*) as count 
                FROM pagos 
                WHERE estado IN ('retraso_leve', 'retraso_grave', 'impago')
            """)
            g.pagos_pendientes = cursor.fetchone()['count']
            
            # Total de alertas
            g.alertas_count = g.clientes_riesgo_count + g.pagos_pendientes
            
        except Exception as e:
            print(f"Error cargando variables globales: {e}")
            # Valores por defecto
            g.total_clientes = 0
            g.clientes_riesgo_count = 0
            g.pagos_pendientes = 0
            g.alertas_count = 0
        finally:
            conn.close()
    else:
        g.total_clientes = 0
        g.clientes_riesgo_count = 0
        g.pagos_pendientes = 0
        g.alertas_count = 0

def obtener_datos_entrenamiento():
    """Obtener datos históricos para entrenamiento"""
    conn = get_db_connection()
    if not conn:
        return []
    
    try:
        query = """
        SELECT cliente_id, estado 
        FROM pagos 
        ORDER BY cliente_id, año, mes
        """
        df = pd.read_sql(query, conn)
        
        # Procesar datos para Markov
        datos_markov = []
        for cliente_id in df['cliente_id'].unique():
            estados = df[df['cliente_id'] == cliente_id]['estado'].tolist()
            if len(estados) >= 2:
                datos_markov.append(estados)
        
        return datos_markov
        
    except Exception as e:
        print(f"Error obteniendo datos: {e}")
        return []
    finally:
        if conn:
            conn.close()

def generar_grafico_estados(estados_clientes):
    """Generar gráfico de estados de clientes"""
    plt.figure(figsize=(10, 6))
    
    labels = [estado['estado_actual'].title().replace('_', ' ') for estado in estados_clientes]
    valores = [estado['cantidad'] for estado in estados_clientes]
    colores = ['#28a745', '#ffc107', '#fd7e14', '#dc3545'][:len(labels)]
    
    plt.bar(labels, valores, color=colores, alpha=0.8)
    plt.title('Distribución de Estados de Clientes', fontsize=14, fontweight='bold')
    plt.xlabel('Estado Actual')
    plt.ylabel('Cantidad de Clientes')
    
    for i, v in enumerate(valores):
        plt.text(i, v + 0.1, str(v), ha='center', va='bottom', fontweight='bold')
    
    plt.tight_layout()
    
    img = io.BytesIO()
    plt.savefig(img, format='png', dpi=100, bbox_inches='tight')
    img.seek(0)
    graph_url = base64.b64encode(img.getvalue()).decode()
    plt.close()
    
    return graph_url

# Rutas de la aplicación
@app.route('/')
def dashboard():
    """Página principal del dashboard mejorado"""
    conn = get_db_connection()
    if not conn:
        return "Error: No se pudo conectar a la base de datos", 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Estadísticas básicas
        cursor.execute("SELECT COUNT(*) as total FROM clientes WHERE apto_prestamo != 'No'")
        total_clientes = cursor.fetchone()['total']
        
        cursor.execute("SELECT estado_actual, COUNT(*) as cantidad FROM clientes WHERE apto_prestamo != 'No' GROUP BY estado_actual")
        estados_clientes = cursor.fetchall()
        
        # Total de clientes rechazados
        cursor.execute("SELECT COUNT(*) as total FROM clientes WHERE apto_prestamo = 'No'")
        total_rechazados = cursor.fetchone()['total']
        
        # Ingresos del mes (noviembre 2025)
        cursor.execute("""
            SELECT ROUND(SUM(monto), 2) as total 
            FROM pagos 
            WHERE estado = 'al_dia' AND mes = 11 AND año = 2025
        """)
        total_ingresos_mes = cursor.fetchone()['total'] or 0
        
        # Pagos al día este mes
        cursor.execute("""
            SELECT COUNT(*) as count 
            FROM pagos 
            WHERE estado = 'al_dia' AND mes = 11 AND año = 2025
        """)
        pagos_al_dia = cursor.fetchone()['count'] or 0
        
        # Pagos en retraso
        cursor.execute("""
            SELECT COUNT(*) as count, ROUND(SUM(monto), 2) as monto
            FROM pagos 
            WHERE estado IN ('retraso_leve', 'retraso_grave')
        """)
        retraso_data = cursor.fetchone()
        pagos_pendientes = retraso_data['count'] or 0
        pagos_retraso = pagos_pendientes
        
        # Impagos
        cursor.execute("""
            SELECT COUNT(*) as count, ROUND(SUM(monto), 2) as monto
            FROM pagos 
            WHERE estado = 'impago'
        """)
        impago_data = cursor.fetchone()
        pagos_impago = impago_data['count'] or 0
        
        # Cartera morosa (retraso + impago)
        cursor.execute("""
            SELECT ROUND(SUM(monto), 2) as total
            FROM pagos 
            WHERE estado IN ('retraso_leve', 'retraso_grave', 'impago')
        """)
        monto_moroso = cursor.fetchone()['total'] or 0
        
        # Total cartera
        cursor.execute("SELECT ROUND(SUM(monto), 2) as total FROM pagos")
        monto_total_cartera = cursor.fetchone()['total'] or 0
        
        # Clientes con mayor riesgo (sin préstamo aprobado)
        cursor.execute("""
            SELECT c.id, c.nombre, c.email, c.estado_actual,
                   COUNT(p.id) as total_pagos,
                   SUM(CASE WHEN p.estado = 'impago' THEN 1 ELSE 0 END) as total_impagos,
                   ROUND(SUM(CASE WHEN p.estado = 'impago' THEN p.monto ELSE 0 END), 2) as deuda_total
            FROM clientes c
            LEFT JOIN pagos p ON c.id = p.cliente_id
            WHERE c.apto_prestamo != 'No'
            GROUP BY c.id, c.nombre, c.email, c.estado_actual
            HAVING total_impagos > 0 OR c.estado_actual = 'impago'
            ORDER BY total_impagos DESC, deuda_total DESC
            LIMIT 5
        """)
        clientes_riesgo = cursor.fetchall()
        
        # Clientes rechazados (No Prestar Nunca Más)
        cursor.execute("""
            SELECT c.id, c.nombre, c.email, c.telefono, c.estado_actual,
                   c.calificacion_crediticia as puntuacion,
                   SUM(CASE WHEN p.estado = 'impago' THEN p.monto ELSE 0 END) as deuda_total,
                   CASE 
                       WHEN c.estado_actual = 'impago' THEN 'Impago permanente'
                       WHEN c.calificacion_crediticia < 30 THEN 'Muy alto riesgo'
                       WHEN c.estado_actual = 'retraso_grave' THEN 'Moroso crónico'
                       ELSE 'Incobrable'
                   END as razon_rechazo
            FROM clientes c
            LEFT JOIN pagos p ON c.id = p.cliente_id
            WHERE c.apto_prestamo = 'No'
            GROUP BY c.id, c.nombre, c.email, c.telefono, c.estado_actual, c.calificacion_crediticia
            ORDER BY c.calificacion_crediticia ASC
        """)
        clientes_rechazados = cursor.fetchall()
        
        return render_template('dashboard.html', 
                             total_clientes=total_clientes,
                             total_rechazados=total_rechazados,
                             estados_clientes=estados_clientes,
                             total_ingresos_mes=total_ingresos_mes,
                             pagos_al_dia=pagos_al_dia,
                             pagos_pendientes=pagos_pendientes,
                             pagos_retraso=pagos_retraso,
                             pagos_impago=pagos_impago,
                             monto_moroso=monto_moroso,
                             monto_total_cartera=monto_total_cartera,
                             clientes_riesgo=clientes_riesgo,
                             clientes_rechazados=clientes_rechazados)
        
    except Exception as e:
        print(f"Error en dashboard: {e}")
        return f"Error: {e}", 500
    finally:
        conn.close()

@app.route('/clientes')
def clientes():
    """Gestión de clientes"""
    conn = get_db_connection()
    if not conn:
        return "Error de conexión", 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Obtener parámetro de filtro
        estado_filtro = request.args.get('estado', '')
        
        # Construir consulta base
        query = """
            SELECT c.*, 
                   COUNT(p.id) as total_pagos,
                   SUM(CASE WHEN p.estado = 'impago' THEN 1 ELSE 0 END) as total_impagos
            FROM clientes c
            LEFT JOIN pagos p ON c.id = p.cliente_id
        """
        
        if estado_filtro:
            query += " WHERE c.estado_actual = %s"
            cursor.execute(query + " GROUP BY c.id ORDER BY c.nombre", (estado_filtro,))
        else:
            cursor.execute(query + " GROUP BY c.id ORDER BY c.nombre")
        
        clientes = cursor.fetchall()
        return render_template('clientes.html', clientes=clientes)
        
    finally:
        conn.close()

@app.route('/api/clientes', methods=['POST'])
def agregar_cliente():
    """API para agregar cliente"""
    try:
        data = request.get_json()
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Realizar evaluación crediticia
        evaluador = EvaluadorCredito()
        evaluacion = evaluador.evaluar_cliente(
            sueldo=float(data.get('sueldo', 0)),
            otros_ingresos=float(data.get('otros_ingresos', 0)),
            gastos_vivienda=float(data.get('gastos_vivienda', 0)),
            tiene_propiedad=data.get('tiene_propiedad', False),
            valor_propiedad=float(data.get('valor_propiedad', 0)) if data.get('tiene_propiedad') else 0,
            estado_actual=data.get('estado_actual', 'al_dia')
        )
        
        cursor.execute(
            '''INSERT INTO clientes (nombre, email, telefono, direccion, estado_actual, sueldo, otros_ingresos, 
                                     gastos_vivienda, tiene_propiedad, valor_propiedad, apto_prestamo, 
                                     evaluacion_ia, calificacion_crediticia, fecha_evaluacion, notas, monto_solicitado) 
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), %s, %s)''',
            (data['nombre'], 
             data['email'],
             data.get('telefono', ''),
             data.get('direccion', ''),
             data.get('estado_actual', 'al_dia'),
             float(data.get('sueldo', 0)),
             float(data.get('otros_ingresos', 0)),
             float(data.get('gastos_vivienda', 0)),
             data.get('tiene_propiedad', False),
             float(data.get('valor_propiedad', 0)) if data.get('tiene_propiedad') else 0,
             evaluacion['apto_prestamo'],
             str(evaluacion),
             evaluacion['puntuacion_final'],
             data.get('notas', ''),
             float(data.get('monto_solicitado', 0)))
        )
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'mensaje': 'Cliente agregado correctamente'})
    except Exception as e:
        return jsonify({'success': False, 'mensaje': str(e)})

@app.route('/api/clientes/<int:cliente_id>', methods=['DELETE'])
def eliminar_cliente(cliente_id):
    """API para eliminar cliente"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Primero eliminar los pagos del cliente
        cursor.execute('DELETE FROM pagos WHERE cliente_id = %s', (cliente_id,))
        # Luego eliminar el cliente
        cursor.execute('DELETE FROM clientes WHERE id = %s', (cliente_id,))
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'mensaje': 'Cliente eliminado correctamente'})
    except Exception as e:
        return jsonify({'success': False, 'mensaje': str(e)})

@app.route('/api/clientes/<int:cliente_id>', methods=['GET'])
def obtener_cliente(cliente_id):
    """API para obtener datos de un cliente específico"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("SELECT * FROM clientes WHERE id = %s", (cliente_id,))
        cliente = cursor.fetchone()
        conn.close()
        
        if cliente:
            return jsonify({'success': True, 'data': cliente})
        else:
            return jsonify({'success': False, 'mensaje': 'Cliente no encontrado'})
    except Exception as e:
        return jsonify({'success': False, 'mensaje': str(e)})

@app.route('/api/clientes/<int:cliente_id>', methods=['PUT'])
def actualizar_cliente(cliente_id):
    """API para actualizar cliente"""
    try:
        data = request.get_json()
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Realizar evaluación crediticia
        evaluador = EvaluadorCredito()
        evaluacion = evaluador.evaluar_cliente(
            sueldo=float(data.get('sueldo', 0)),
            otros_ingresos=float(data.get('otros_ingresos', 0)),
            gastos_vivienda=float(data.get('gastos_vivienda', 0)),
            tiene_propiedad=data.get('tiene_propiedad', False),
            valor_propiedad=float(data.get('valor_propiedad', 0)) if data.get('tiene_propiedad') else 0,
            estado_actual=data.get('estado_actual', 'al_dia')
        )
        
        cursor.execute(
            '''UPDATE clientes SET nombre=%s, email=%s, telefono=%s, direccion=%s, estado_actual=%s, sueldo=%s, 
                                   otros_ingresos=%s, gastos_vivienda=%s, tiene_propiedad=%s, 
                                   valor_propiedad=%s, apto_prestamo=%s, evaluacion_ia=%s, 
                                   calificacion_crediticia=%s, fecha_evaluacion=NOW(), notas=%s, monto_solicitado=%s
               WHERE id=%s''',
            (data['nombre'], 
             data['email'],
             data.get('telefono', ''),
             data.get('direccion', ''),
             data['estado_actual'],
             float(data.get('sueldo', 0)),
             float(data.get('otros_ingresos', 0)),
             float(data.get('gastos_vivienda', 0)),
             data.get('tiene_propiedad', False),
             float(data.get('valor_propiedad', 0)) if data.get('tiene_propiedad') else 0,
             evaluacion['apto_prestamo'],
             str(evaluacion),
             evaluacion['puntuacion_final'],
             data.get('notas', ''),
             float(data.get('monto_solicitado', 0)),
             cliente_id)
        )
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'mensaje': 'Cliente actualizado correctamente'})
    except Exception as e:
        return jsonify({'success': False, 'mensaje': str(e)})

@app.route('/cliente/<int:cliente_id>/pagos')
def pagos_cliente(cliente_id):
    """Página de pagos específicos de un cliente"""
    conn = get_db_connection()
    if not conn:
        return "Error de conexión", 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Obtener información del cliente
        cursor.execute("SELECT * FROM clientes WHERE id = %s", (cliente_id,))
        cliente = cursor.fetchone()
        
        if not cliente:
            return "Cliente no encontrado", 404
        
        # Obtener todos los pagos del cliente ordenados por año y mes
        cursor.execute("""
            SELECT p.*, 
                   CASE 
                     WHEN p.estado = 'al_dia' THEN 'Al día'
                     WHEN p.estado = 'retraso_leve' THEN 'Retraso leve'
                     WHEN p.estado = 'retraso_grave' THEN 'Retraso grave'
                     ELSE 'Impago'
                   END as estado_display
            FROM pagos p 
            WHERE p.cliente_id = %s 
            ORDER BY p.año DESC, p.mes DESC
        """, (cliente_id,))
        pagos = cursor.fetchall()
        
        # Agrupar pagos por año
        pagos_por_anio = {}
        for pago in pagos:
            anio = pago['año']
            if anio not in pagos_por_anio:
                pagos_por_anio[anio] = []
            pagos_por_anio[anio].append(pago)
        
        # Calcular estadísticas del cliente
        cursor.execute("""
            SELECT 
                COUNT(*) as total_pagos,
                SUM(CASE WHEN estado = 'al_dia' THEN 1 ELSE 0 END) as pagos_al_dia,
                SUM(CASE WHEN estado = 'impago' THEN 1 ELSE 0 END) as pagos_impago,
                ROUND(SUM(monto), 2) as total_pagado,
                ROUND(SUM(CASE WHEN estado = 'impago' THEN monto ELSE 0 END), 2) as deuda_total,
                ROUND(AVG(monto), 2) as promedio_mensual
            FROM pagos 
            WHERE cliente_id = %s
        """, (cliente_id,))
        estadisticas = cursor.fetchone()
        
        return render_template('pagos_cliente.html', 
                             cliente=cliente, 
                             pagos_por_anio=pagos_por_anio,
                             estadisticas=estadisticas)
        
    except Exception as e:
        return f"Error: {e}", 500
    finally:
        conn.close()

@app.route('/api/cliente/<int:cliente_id>/pagos', methods=['POST'])
def agregar_pago_cliente(cliente_id):
    """API para agregar pago a un cliente específico"""
    try:
        data = request.get_json()
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Verificar que el cliente existe
        cursor.execute("SELECT id FROM clientes WHERE id = %s", (cliente_id,))
        if not cursor.fetchone():
            return jsonify({'success': False, 'mensaje': 'Cliente no encontrado'})
        
        cursor.execute(
            'INSERT INTO pagos (cliente_id, mes, año, estado, monto, descripcion) VALUES (%s, %s, %s, %s, %s, %s)',
            (cliente_id, data['mes'], data['año'], data['estado'], data['monto'], data.get('descripcion', ''))
        )
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'mensaje': 'Pago agregado correctamente'})
    except Exception as e:
        return jsonify({'success': False, 'mensaje': str(e)})

@app.route('/api/pagos/<int:pago_id>/estado', methods=['PUT'])
def actualizar_estado_pago(pago_id):
    """API para actualizar solo el estado de un pago"""
    try:
        data = request.get_json()
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute(
            'UPDATE pagos SET estado = %s WHERE id = %s',
            (data['estado'], pago_id)
        )
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'mensaje': 'Estado del pago actualizado correctamente'})
    except Exception as e:
        return jsonify({'success': False, 'mensaje': str(e)})

@app.route('/pagos')
def pagos():
    """Página de gestión de pagos organizada por clientes"""
    conn = get_db_connection()
    if not conn:
        return "Error de conexión", 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Obtener parámetro de filtro
        estado_filtro = request.args.get('estado', '')
        
        # Construir consulta base
        query = """
            SELECT p.*, c.nombre as cliente_nombre, c.monto_solicitado
            FROM pagos p 
            JOIN clientes c ON p.cliente_id = c.id 
        """
        
        if estado_filtro:
            query += " WHERE p.estado = %s"
            cursor.execute(query + " ORDER BY c.nombre, p.año DESC, p.mes DESC", (estado_filtro,))
        else:
            cursor.execute(query + " ORDER BY c.nombre, p.año DESC, p.mes DESC")
        
        pagos_data = cursor.fetchall()
        
        # Fecha actual para cálculos
        fecha_hoy = datetime(2025, 11, 10)
        
        # Agrupar pagos por cliente
        pagos_por_cliente = {}
        for pago in pagos_data:
            cliente_id = pago['cliente_id']
            if cliente_id not in pagos_por_cliente:
                pagos_por_cliente[cliente_id] = {
                    'cliente_id': cliente_id,
                    'cliente_nombre': pago['cliente_nombre'],
                    'pagos': [],
                    'total_pagos': 0,
                    'pagos_al_dia': 0,
                    'pagos_impago': 0,
                    'pagos_en_retraso': 0,
                    'deuda_total': 0,
                    'proximo_pago': None,
                    'proximo_pago_fecha': None,
                    'proximo_pago_monto': 0,
                    'pagos_hechos': 0,
                    'pagos_faltantes': 0,
                    'fecha_termina_prestamo': None,
                    'deuda_hoy': 0
                }
            
            pagos_por_cliente[cliente_id]['pagos'].append(pago)
            pagos_por_cliente[cliente_id]['total_pagos'] += 1
            
            if pago['estado'] == 'al_dia':
                pagos_por_cliente[cliente_id]['pagos_al_dia'] += 1
                pagos_por_cliente[cliente_id]['pagos_hechos'] += 1
            elif pago['estado'] in ['retraso_leve', 'retraso_grave']:
                pagos_por_cliente[cliente_id]['pagos_en_retraso'] += 1
                pagos_por_cliente[cliente_id]['pagos_hechos'] += 1
            elif pago['estado'] == 'impago':
                pagos_por_cliente[cliente_id]['pagos_impago'] += 1
                pagos_por_cliente[cliente_id]['deuda_total'] += pago['monto']
                pagos_por_cliente[cliente_id]['deuda_hoy'] += pago['monto']
        
        # Obtener información del préstamo para cada cliente
        cursor.execute("""
            SELECT cliente_id, 
                   MIN(mes) as mes_inicio, 
                   MIN(año) as año_inicio,
                   MAX(mes) as mes_fin, 
                   MAX(año) as año_fin,
                   COUNT(*) as total_cuotas_previstas,
                   AVG(monto) as cuota_promedio
            FROM pagos 
            GROUP BY cliente_id
        """)
        info_prestamos = {row['cliente_id']: row for row in cursor.fetchall()}
        
        # Calcular información detallada para cada cliente
        for cliente_id, info_cliente in pagos_por_cliente.items():
            if info_cliente['pagos']:
                ultimo_pago = info_cliente['pagos'][0]  # El primero después del ORDER BY DESC
                
                # Calcular próximo pago
                mes_proximo = ultimo_pago['mes'] + 1
                año_proximo = ultimo_pago['año']
                
                if mes_proximo > 12:
                    mes_proximo = 1
                    año_proximo += 1
                
                # Crear fecha estimada del próximo pago
                try:
                    fecha_proximo = datetime(año_proximo, mes_proximo, 15)
                    mes_nombre = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic'][mes_proximo-1]
                    info_cliente['proximo_pago'] = f"{mes_nombre}/{año_proximo}"
                    info_cliente['proximo_pago_fecha'] = fecha_proximo
                    info_cliente['proximo_pago_monto'] = ultimo_pago['monto']
                except:
                    info_cliente['proximo_pago'] = 'No disponible'
                
                # Obtener información del préstamo
                if cliente_id in info_prestamos:
                    prestamo = info_prestamos[cliente_id]
                    
                    # Pagos faltantes
                    info_cliente['pagos_faltantes'] = prestamo['total_cuotas_previstas'] - info_cliente['pagos_hechos']
                    
                    # Fecha cuando paga todo (mes/año final del préstamo)
                    try:
                        mes_fin = prestamo['mes_fin']
                        año_fin = prestamo['año_fin']
                        mes_nombre_fin = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic'][mes_fin-1]
                        info_cliente['fecha_termina_prestamo'] = f"{mes_nombre_fin} {año_fin}"
                        fecha_termina = datetime(año_fin, mes_fin, 15)
                        info_cliente['fecha_termina_timestamp'] = fecha_termina
                    except:
                        info_cliente['fecha_termina_prestamo'] = 'No disponible'
                    
                    # Calcular deuda al día de hoy (pagos impagos hasta hoy)
                    deuda_hoy = 0
                    for pago in info_cliente['pagos']:
                        try:
                            fecha_pago = datetime(pago['año'], pago['mes'], 15)
                            if fecha_pago <= fecha_hoy and pago['estado'] == 'impago':
                                deuda_hoy += pago['monto']
                        except:
                            pass
                    info_cliente['deuda_hoy'] = deuda_hoy
        
        # Convertir a lista para la template
        pagos_por_cliente_lista = list(pagos_por_cliente.values())
        
        # Calcular totales por estado para los filtros
        cursor.execute("SELECT estado, COUNT(*) as total FROM pagos GROUP BY estado")
        estados_totales = cursor.fetchall()
        
        total_al_dia = next((item['total'] for item in estados_totales if item['estado'] == 'al_dia'), 0)
        total_retraso_leve = next((item['total'] for item in estados_totales if item['estado'] == 'retraso_leve'), 0)
        total_retraso_grave = next((item['total'] for item in estados_totales if item['estado'] == 'retraso_grave'), 0)
        total_impago = next((item['total'] for item in estados_totales if item['estado'] == 'impago'), 0)
        
        # Obtener clientes para el formulario
        cursor.execute("SELECT id, nombre FROM clientes ORDER BY nombre")
        clientes = cursor.fetchall()
        
        return render_template('pagos.html', 
                             pagos=pagos_data,
                             pagos_por_cliente=pagos_por_cliente_lista,
                             clientes=clientes,
                             total_al_dia=total_al_dia,
                             total_retraso_leve=total_retraso_leve,
                             total_retraso_grave=total_retraso_grave,
                             total_impago=total_impago,
                             fecha_hoy=fecha_hoy.strftime('%d de %B de %Y'))
        
    finally:
        conn.close()

@app.route('/api/pagos', methods=['POST'])
def agregar_pago():
    """API para agregar pago"""
    try:
        data = request.get_json()
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute(
            'INSERT INTO pagos (cliente_id, mes, año, estado, monto, descripcion) VALUES (%s, %s, %s, %s, %s, %s)',
            (data['cliente_id'], data['mes'], data['año'], data['estado'], data['monto'], data.get('descripcion', ''))
        )
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'mensaje': 'Pago agregado correctamente'})
    except Exception as e:
        return jsonify({'success': False, 'mensaje': str(e)})

@app.route('/api/pagos/<int:pago_id>', methods=['PUT'])
def actualizar_pago(pago_id):
    """API para actualizar pago"""
    try:
        data = request.get_json()
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute(
            'UPDATE pagos SET cliente_id=%s, mes=%s, año=%s, estado=%s, monto=%s, descripcion=%s WHERE id=%s',
            (data['cliente_id'], data['mes'], data['año'], data['estado'], data['monto'], data.get('descripcion', ''), pago_id)
        )
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'mensaje': 'Pago actualizado correctamente'})
    except Exception as e:
        return jsonify({'success': False, 'mensaje': str(e)})

@app.route('/api/pagos/<int:pago_id>', methods=['DELETE'])
def eliminar_pago(pago_id):
    """API para eliminar pago"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM pagos WHERE id = %s', (pago_id,))
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'mensaje': 'Pago eliminado correctamente'})
    except Exception as e:
        return jsonify({'success': False, 'mensaje': str(e)})

@app.route('/api/evaluar-crediticio', methods=['POST'])
def evaluar_crediticio():
    """API para evaluar la aptitud crediticia de un cliente"""
    try:
        data = request.get_json()
        
        # Validar datos requeridos
        campos_requeridos = ['sueldo', 'gastos_vivienda']
        for campo in campos_requeridos:
            if campo not in data:
                return jsonify({'success': False, 'mensaje': f'Falta el campo {campo}'}), 400
        
        # Obtener información del cliente
        cliente_id = data.get('cliente_id')
        sueldo = float(data.get('sueldo', 0))
        otros_ingresos = float(data.get('otros_ingresos', 0))
        gastos_vivienda = float(data.get('gastos_vivienda', 0))
        tiene_propiedad = data.get('tiene_propiedad', False)
        valor_propiedad = float(data.get('valor_propiedad', 0)) if tiene_propiedad else 0
        
        # Obtener estado actual del cliente
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        estado_actual = 'pendiente'
        if cliente_id:
            cursor.execute("SELECT estado_actual FROM clientes WHERE id = %s", (cliente_id,))
            cliente = cursor.fetchone()
            if cliente:
                estado_actual = cliente['estado_actual']
        
        # Evaluar crediticio
        evaluador = EvaluadorCredito()
        evaluacion = evaluador.evaluar_cliente(
            sueldo=sueldo,
            otros_ingresos=otros_ingresos,
            gastos_vivienda=gastos_vivienda,
            tiene_propiedad=tiene_propiedad,
            valor_propiedad=valor_propiedad,
            estado_actual=estado_actual
        )
        
        conn.close()
        
        return jsonify({
            'success': True,
            'evaluacion': evaluacion
        })
    
    except Exception as e:
        return jsonify({'success': False, 'mensaje': str(e)}), 500

@app.route('/predicciones')
def predicciones():
    """Página de predicciones IA"""
    conn = get_db_connection()
    if not conn:
        return "Error de conexión", 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM clientes ORDER BY nombre")
        clientes = cursor.fetchall()
        return render_template('predicciones.html', clientes=clientes)
    finally:
        conn.close()

@app.route('/api/predecir/<int:cliente_id>')
def predecir_cliente(cliente_id):
    """Realizar predicción para un cliente"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Obtener cliente
        cursor.execute("SELECT * FROM clientes WHERE id = %s", (cliente_id,))
        cliente = cursor.fetchone()
        
        # Obtener último estado
        cursor.execute("""
            SELECT estado FROM pagos 
            WHERE cliente_id = %s 
            ORDER BY año DESC, mes DESC 
            LIMIT 1
        """, (cliente_id,))
        ultimo_pago = cursor.fetchone()
        
        conn.close()
        
        if not ultimo_pago:
            return jsonify({'error': 'No hay datos del cliente'})
        
        estado_actual = ultimo_pago['estado']
        
        # Entrenar y predecir con Markov
        datos = obtener_datos_entrenamiento()
        markov = CadenaMarkovPagos()
        markov.entrenar(datos)
        predicciones, probabilidades = markov.predecir(estado_actual, 3)
        nivel_riesgo = markov.calcular_riesgo(estado_actual)

        # Evaluación financiera (SAFA)
        try:
            evaluador = EvaluadorCredito()
            sueldo = float(cliente.get('sueldo', 0) or 0)
            otros_ingresos = float(cliente.get('otros_ingresos', 0) or 0)
            gastos_vivienda = float(cliente.get('gastos_vivienda', 0) or 0)
            tiene_propiedad = bool(cliente.get('tiene_propiedad'))
            valor_propiedad = float(cliente.get('valor_propiedad', 0) or 0)

            evaluacion = evaluador.evaluar_cliente(
                sueldo=sueldo,
                otros_ingresos=otros_ingresos,
                gastos_vivienda=gastos_vivienda,
                tiene_propiedad=tiene_propiedad,
                valor_propiedad=valor_propiedad,
                estado_actual=estado_actual
            )
        except Exception as e:
            evaluacion = {'error': str(e)}

        return jsonify({
            'success': True,
            'cliente': cliente,
            'estado_actual': estado_actual,
            'nivel_riesgo': nivel_riesgo,
            'predicciones': predicciones,
            'probabilidades': probabilidades,
            'evaluacion': evaluacion
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# Nuevas rutas para reportes
@app.route('/reportes')
def reportes():
    """Página principal de reportes"""
    return render_template('reportes.html')

@app.route('/reportes/clientes')
def reporte_clientes():
    """Reporte detallado de clientes"""
    conn = get_db_connection()
    if not conn:
        return "Error de conexión", 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT c.*, 
                   COUNT(p.id) as total_pagos,
                   SUM(CASE WHEN p.estado = 'impago' THEN 1 ELSE 0 END) as total_impagos,
                   ROUND(SUM(CASE WHEN p.estado = 'impago' THEN p.monto ELSE 0 END), 2) as deuda_total
            FROM clientes c
            LEFT JOIN pagos p ON c.id = p.cliente_id
            GROUP BY c.id
            ORDER BY deuda_total DESC
        """)
        clientes = cursor.fetchall()
        return render_template('reporte_clientes.html', clientes=clientes)
    finally:
        conn.close()

@app.route('/reportes/pagos')
def reporte_pagos():
    """Reporte detallado de pagos"""
    conn = get_db_connection()
    if not conn:
        return "Error de conexión", 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT p.*, c.nombre as cliente_nombre 
            FROM pagos p 
            JOIN clientes c ON p.cliente_id = c.id 
            ORDER BY p.año DESC, p.mes DESC
        """)
        pagos = cursor.fetchall()
        return render_template('reporte_pagos.html', pagos=pagos)
    finally:
        conn.close()


def generar_excel_pagos():
    """Genera un Excel detallado con todos los pagos (cliente, monto, fecha, estado, descripción)."""
    conn = get_db_connection()
    if not conn:
        raise RuntimeError('No se pudo conectar a la base de datos')

    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT p.id as pago_id, p.cliente_id, p.mes, p.año, p.monto, p.estado, p.descripcion, p.fecha_pago,
                   c.nombre as cliente_nombre, c.email as cliente_email, c.telefono as cliente_telefono
            FROM pagos p
            JOIN clientes c ON p.cliente_id = c.id
            ORDER BY p.año DESC, p.mes DESC, c.nombre
        """)

        pagos = cursor.fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Pagos Detallados'

        # Estilos
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        align_center = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = [
            'Pago ID', 'Cliente ID', 'Nombre Cliente', 'Email', 'Teléfono', 'Mes', 'Año', 'Fecha Pago', 'Monto', 'Estado', 'Descripción'
        ]

        for col_idx, title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border

        # Rellenar filas
        fila = 2
        for p in pagos:
            ws.cell(row=fila, column=1, value=p.get('pago_id'))
            ws.cell(row=fila, column=2, value=p.get('cliente_id'))
            ws.cell(row=fila, column=3, value=p.get('cliente_nombre'))
            ws.cell(row=fila, column=4, value=p.get('cliente_email'))
            ws.cell(row=fila, column=5, value=p.get('cliente_telefono'))
            ws.cell(row=fila, column=6, value=p.get('mes'))
            ws.cell(row=fila, column=7, value=p.get('año'))

            # Fecha de pago: si el campo fecha_pago existe (TIMESTAMP), usarlo; si no, crear con día 15
            fecha_pago = p.get('fecha_pago')
            if fecha_pago is None:
                try:
                    fecha_pago = datetime(p.get('año') or 1, p.get('mes') or 1, 15)
                except Exception:
                    fecha_pago = None

            cell_fecha = ws.cell(row=fila, column=8, value=fecha_pago)
            if fecha_pago:
                cell_fecha.number_format = 'DD/MM/YYYY'

            cell_monto = ws.cell(row=fila, column=9, value=float(p.get('monto') or 0))
            cell_monto.number_format = '€#,##0.00'

            ws.cell(row=fila, column=10, value=p.get('estado'))
            ws.cell(row=fila, column=11, value=p.get('descripcion'))

            # Aplicar estilo de borde a toda la fila
            for col in range(1, len(headers) + 1):
                c = ws.cell(row=fila, column=col)
                c.border = border
                if col in (1,2,6,7):
                    c.alignment = Alignment(horizontal='center')

            fila += 1

        # Ajustar anchos
        widths = [10, 10, 30, 30, 15, 8, 8, 15, 12, 15, 40]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # Guardar en buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer

    finally:
        cursor.close()
        conn.close()


@app.route('/api/descargar-informe-pagos')
def descargar_informe_pagos():
    """Endpoint para descargar informe de pagos en Excel con detalles (nombre, monto, fecha)."""
    try:
        excel_buffer = generar_excel_pagos()
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'INFORME_PAGOS_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Rutas de configuración y perfil
@app.route('/configuracion')
def configuracion():
    """Página de configuración"""
    return render_template('configuracion.html')

@app.route('/perfil')
def perfil():
    """Página de perfil de usuario"""
    return render_template('perfil.html')

@app.route('/ayuda')
def ayuda():
    """Página de ayuda"""
    return render_template('ayuda.html')

def generar_excel_librerias():
    """Genera un archivo Excel con análisis de librerías"""
    
    # Datos de librerías analizadas
    librerias_info = {
        'Flask': {
            'version': '3.1.2',
            'usado_en': ['app.py'],
            'uso': 'Framework web (rutas, templates, request/response)',
            'estado': '✓ Necesaria',
            'riesgo': 'Bajo'
        },
        'pandas': {
            'version': '2.3.3',
            'usado_en': ['app.py'],
            'uso': 'Análisis de datos, lectura SQL a DataFrame',
            'estado': '✓ Necesaria',
            'riesgo': 'Bajo'
        },
        'numpy': {
            'version': '2.3.4',
            'usado_en': ['modelos_ia.py'],
            'uso': 'Importado - cálculos numéricos',
            'estado': '⚠ Opcional',
            'riesgo': 'Bajo'
        },
        'scikit-learn': {
            'version': '1.7.2',
            'usado_en': ['modelos_ia.py'],
            'uso': 'Machine Learning (RandomForest)',
            'estado': '✓ Necesaria',
            'riesgo': 'Bajo'
        },
        'mysql-connector-python': {
            'version': '9.5.0',
            'usado_en': ['database.py'],
            'uso': 'Conexión a base de datos MySQL',
            'estado': '✓ Necesaria',
            'riesgo': 'Bajo'
        },
        'matplotlib': {
            'version': '3.10.7',
            'usado_en': ['app.py'],
            'uso': 'Generación de gráficos PNG',
            'estado': '✓ Necesaria',
            'riesgo': 'Bajo'
        },
        'python-dotenv': {
            'version': '1.2.1',
            'usado_en': ['app.py', 'database.py'],
            'uso': 'Cargar variables de entorno',
            'estado': '✓ Necesaria',
            'riesgo': 'Bajo'
        }
    }
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Hoja 1: Resumen
    ws = wb.create_sheet('📋 Resumen', 0)
    
    ws['A1'] = 'ANÁLISIS DE LIBRERÍAS - SISTEMA CONTADOR'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['A2'] = f'Generado: {datetime.now().strftime("%d de %B de %Y, %H:%M")}'
    ws['A2'].font = Font(italic=True, size=10)
    
    fila = 4
    metricas = [
        ('Total librerías', '7'),
        ('Librerías necesarias', '7'),
        ('Librerías opcionales', '1'),
        ('Estado', '✓ Operativo'),
        ('Warnings', '0/8 ✓ Resueltos')
    ]
    
    for label, valor in metricas:
        ws[f'A{fila}'] = label
        ws[f'B{fila}'] = valor
        ws[f'B{fila}'].font = Font(bold=True)
        fila += 1
    
    # Hoja 2: Librerías detalladas
    ws_libs = wb.create_sheet('📦 Librerías', 1)
    
    headers = ['Librería', 'Versión', 'Usado en', 'Descripción', 'Estado']
    for col, header in enumerate(headers, 1):
        cell = ws_libs.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    fila = 2
    for lib_name, lib_data in librerias_info.items():
        ws_libs.cell(row=fila, column=1).value = lib_name
        ws_libs.cell(row=fila, column=2).value = lib_data['version']
        ws_libs.cell(row=fila, column=3).value = ', '.join(lib_data['usado_en'])
        ws_libs.cell(row=fila, column=4).value = lib_data['uso']
        ws_libs.cell(row=fila, column=5).value = lib_data['estado']
        
        if '✓' in lib_data['estado']:
            ws_libs.cell(row=fila, column=5).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        else:
            ws_libs.cell(row=fila, column=5).fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        
        fila += 1
    
    ws_libs.column_dimensions['A'].width = 20
    ws_libs.column_dimensions['B'].width = 12
    ws_libs.column_dimensions['C'].width = 20
    ws_libs.column_dimensions['D'].width = 40
    ws_libs.column_dimensions['E'].width = 15
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    
    # Guardar en buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

@app.route('/api/descargar-informe-librerias')
def descargar_informe_librerias():
    """API para descargar informe de librerías en Excel"""
    try:
        excel_buffer = generar_excel_librerias()
        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'INFORME_LIBRERIAS_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def abrir_navegador():
    """Abrir navegador automáticamente después de 2 segundos"""
    time.sleep(2)
    webbrowser.open('http://localhost:5000')

if __name__ == '__main__':
    # Inicializar base de datos
    if init_database():
        print("✅ Sistema iniciado correctamente")
        print("🌐 http://localhost:5000")
        print("🖥️  Abriendo navegador automáticamente...")
        
        # Abrir navegador en hilo separado
        threading.Thread(target=abrir_navegador).start()
        
        # Ejecutar servidor
        app.run(debug=True, host='0.0.0.0', port=5000)
    else:
        print("❌ Error al iniciar la base de datos")